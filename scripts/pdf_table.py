import os
from concurrent.futures import ThreadPoolExecutor
import cv2
import pytesseract
import pandas as pd
import numpy as np
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pdf2image import convert_from_path
from functools import partial

# Options
screen_width = 1600
dpi = 1200
concurrent_ocr = False  # Set to False to disable parallel processing

# Global variable
scale = 0


def preprocess_table_cell(cell_img, empty_thresh=0.002):
    """
    Preprocess a table cell image for OCR.

    Steps:
    1. Binarize the grayscale cell image
    2. Remove connected components touching the border (table lines / artifacts)
    3. Skip if cell is mostly empty

    Parameters:
        cell_img: np.array, grayscale cell image
        empty_thresh: fraction of non-background pixels to consider non-empty

    Returns:
        processed_img: np.array ready for OCR, or None if the cell is empty
    """

    # Binarize using an Otsu threshold
    _, binary = cv2.threshold(cell_img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if binary is None:
        return None

    # Invert: text should be white (255), background black (0)
    binary = 255 - binary

    # Remove connected components touching the border
    h, w = binary.shape
    mask = np.zeros((h + 2, w + 2), np.uint8)
    cv2.floodFill(binary, mask, (0, 0), 0)
    cv2.floodFill(binary, mask, (w - 1, 0), 0)
    cv2.floodFill(binary, mask, (0, h - 1), 0)
    cv2.floodFill(binary, mask, (w - 1, h - 1), 0)

    # Check if the cell is essentially empty
    non_zero_fraction = np.count_nonzero(binary) / (binary.shape[0] * binary.shape[1])
    if non_zero_fraction < empty_thresh:
        return None  # empty cell

    # Optional: slight dilation to restore thin text
    kernel = np.ones((1, 1), np.uint8)
    processed_img = cv2.dilate(binary, kernel, iterations=1)
    if processed_img is None:
        return None

    # Invert so text is black on white
    binary_for_ocr = 255 - processed_img

    return binary_for_ocr


def ocr_cell(cell_img, lang="deu", preview_scale=0.3):
    ocr_result = ""

    processed = preprocess_table_cell(cell_img)
    if processed is not None:
        ocr_result = pytesseract.image_to_string(processed, config='--oem 3 --psm 6', lang=lang).strip()
        print(f"Cell text:\n{ocr_result}")

        # --- View cell OCR ---
        if not concurrent_ocr:
            preview = cv2.resize(
                processed,
                None,
                fx=preview_scale,
                fy=preview_scale,
                interpolation=cv2.INTER_AREA
            )
            cv2.imshow("OCR Preview", preview)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
    else:
        print(f"Cell is empty")

    return ocr_result


def render_table_grid(image, col_x, row_y):
    grid = image.copy()

    # Draw columns between the first and last row
    y_start, y_end = row_y[0], row_y[-1]
    for x in col_x:
        cv2.line(grid, (x, y_start), (x, y_end), (0, 255, 0), 2)

    # Draw rows between the first and last column
    x_start, x_end = col_x[0], col_x[-1]
    for y in row_y:
        cv2.line(grid, (x_start, y), (x_end, y), (255, 0, 0), 2)

    return grid


def save_selection(filename, col_x, row_y, merged_cells, rotated_cells, image_shape, dpi):
    data = {
        "dpi": dpi,  # ✅ NEW
        "col_x": [int(x) for x in col_x],
        "row_y": [int(y) for y in row_y],
        "merged_cells": [[int(v) for v in cell] for cell in merged_cells],
        "rotated_cells": [[int(v) for v in cell] for cell in rotated_cells],
        "image_shape": [int(s) for s in image_shape]
    }
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def load_selection(filename):
    with open(filename, "r", encoding="utf-8") as f:
        data = json.load(f)

    saved_dpi = data.get("dpi", None)  # backward compatible

    return (
        data["col_x"],
        data["row_y"],
        data["merged_cells"],
        data.get("rotated_cells", []),
        tuple(data["image_shape"]),
        saved_dpi
    )



def rotate_for_ocr(img, angle):
    if angle == 90:
        return cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
    elif angle == 270:
        return cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
    return img


def intersects_merged_cell(x1, y1, x2, y2, merged_cells, col_x, row_y):
    for r1, c1, r2, c2 in merged_cells:
        mx1, mx2 = col_x[c1], col_x[c2 + 1]
        my1, my2 = row_y[r1], row_y[r2 + 1]

        # vertical line
        if x1 == x2:
            if mx1 < x1 < mx2 and not (y2 <= my1 or y1 >= my2):
                return True

        # horizontal line
        if y1 == y2:
            if my1 < y1 < my2 and not (x2 <= mx1 or x1 >= mx2):
                return True
    return False


def collect_lines(image, axis="x", title="Click borders, ENTER to finish", initial_lines=None):
    img = image.copy()
    lines = initial_lines.copy() if initial_lines else []

    def on_click(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            orig_x = int(x / scale)
            orig_y = int(y / scale)
            if axis == "x":
                lines.append(orig_x)
            else:
                lines.append(orig_y)
        elif event == cv2.EVENT_RBUTTONDOWN:
            orig_val = int(x / scale) if axis == "x" else int(y / scale)
            if lines:
                # Find the closest line and remove it if it's within a small threshold
                closest = min(lines, key=lambda l: abs(l - orig_val))
                if abs(closest - orig_val) < 20: # threshold in pixels
                    lines.remove(closest)

    cv2.namedWindow(title)
    cv2.setMouseCallback(title, on_click)

    while True:
        display = img.copy()

        for v in lines:
            if axis == "x":
                cv2.line(display, (v, 0), (v, display.shape[0]), (0, 0, 255), 2)
            else:
                cv2.line(display, (0, v), (display.shape[1], v), (255, 0, 0), 2)

        display_img = cv2.resize(
            display,
            None,
            fx=scale,
            fy=scale,
            interpolation=cv2.INTER_AREA
        )

        cv2.imshow(title, display_img)
        key = cv2.waitKey(20) & 0xFF

        # ENTER → finish
        if key == 13:
            break

        # BACKSPACE or 'z' → undo
        if key in (8, ord('z')):
            if lines:
                lines.pop()

        # ESC → remove all lines
        if key == 27:
            lines.clear()

    cv2.destroyAllWindows()
    return sorted(lines)


def ocr_worker(task, rotated_lookup, ocr_lang):
    r, c, img = task

    # rotate if needed
    if (r, c) in rotated_lookup:
        img = rotate_for_ocr(img, rotated_lookup[(r, c)])

    text = ocr_cell(
        img,
        lang=ocr_lang  # never show windows in threads
    )
    return r, c, text


def find_merged_region(r, c, merged_cells):
    """
    Returns (r1, c1, r2, c2) if (r,c) is inside a merged cell,
    otherwise None
    """
    for r1, c1, r2, c2 in merged_cells:
        if r1 <= r <= r2 and c1 <= c <= c2:
            return r1, c1, r2, c2
    return None


def apply_excel_layout(
    xlsx_path,
    merged_cells,
    rotated_cells
):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # ------------------------------
    # Apply merged cells
    # ------------------------------
    for r1, c1, r2, c2 in merged_cells:
        # Excel is 1-based
        ws.merge_cells(
            start_row=r1 + 1,
            start_column=c1 + 1,
            end_row=r2 + 1,
            end_column=c2 + 1
        )

        # center text in merged cell
        cell = ws.cell(row=r1 + 1, column=c1 + 1)
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center",
            wrap_text=True
        )

    # ------------------------------
    # Apply rotated text
    # ------------------------------
    for r1, c1, r2, c2, angle in rotated_cells:
        cell = ws.cell(row=r1 + 1, column=c1 + 1)

        cell.alignment = Alignment(
            textRotation=angle,
            vertical="center",
            horizontal="center",
            wrap_text=True
        )

    wb.save(xlsx_path)


def createTable():
    print('\nWhich PDF table file to convert to .csv?')
    files = []
    for file in os.listdir("input/"):
        if file.endswith(".pdf"):
            print(f" {str(len(files))} --- {file}")
            files.append(file)

    selection = int(input("Select a file to use: "))
    path = "input/" + files[selection]
    page_number = 0
    ocr_lang = "deu"

    savefile = f"output/{files[selection]}_selection.json"

    # ==============================
    # PDF → IMAGE
    # ==============================
    images = convert_from_path(path, dpi=dpi)
    image = np.array(images[page_number])
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    global scale
    scale = screen_width / image.shape[1]

    col_x, row_y, merged_cells, rotated_cells = [], [], [], []

    # ==============================
    # Load previous selection if exists
    # ==============================
    if os.path.exists(savefile):
        print("\nPrevious selection found. Loading...")
        col_x, row_y, merged_cells, rotated_cells, old_shape, saved_dpi = load_selection(savefile)

        if saved_dpi and saved_dpi != dpi:
            scale_factor = dpi / saved_dpi
            print(f"Rescaling selection from DPI {saved_dpi} → {dpi}")

            col_x = [int(x * scale_factor) for x in col_x]
            row_y = [int(y * scale_factor) for y in row_y]

        edit = input("Edit previous selection? (y/n): ").lower() == 'y'
    else:
        edit = True

    if edit:
        col_x = collect_lines(image, axis="x", title="Select columns", initial_lines=col_x)
        row_y = collect_lines(image, axis="y", title="Select rows", initial_lines=row_y)

        # ==============================
        # Merge cells
        # ==============================
        clicks = []

        def merge_click(event, x, y, flags, param):
            orig_x = int(x / scale)
            orig_y = int(y / scale)
            if event == cv2.EVENT_LBUTTONDOWN:
                clicks.append((orig_x, orig_y))
            elif event == cv2.EVENT_RBUTTONDOWN:
                # Find if this click is inside any merged cell
                for i, (r1, c1, r2, c2) in enumerate(merged_cells):
                    x1, x2 = col_x[c1], col_x[c2 + 1]
                    y1, y2 = row_y[r1], row_y[r2 + 1]
                    if x1 <= orig_x <= x2 and y1 <= orig_y <= y2:
                        merged_cells.pop(i)
                        break

        def find_index(val, arr):
            return max(0, np.searchsorted(arr, val) - 1)

        grid_image = render_table_grid(image, col_x, row_y)
        cv2.namedWindow("Merge cells")
        cv2.setMouseCallback("Merge cells", merge_click)

        while True:
            display = image.copy()

            # draw vertical grid lines (clipped)
            for i, x in enumerate(col_x):
                for r in range(len(row_y) - 1):
                    y1, y2 = row_y[r], row_y[r + 1]
                    if not intersects_merged_cell(x, y1, x, y2, merged_cells, col_x, row_y):
                        cv2.line(display, (x, y1), (x, y2), (0, 255, 0), 2)

            # draw horizontal grid lines (clipped)
            for i, y in enumerate(row_y):
                for c in range(len(col_x) - 1):
                    x1, x2 = col_x[c], col_x[c + 1]
                    if not intersects_merged_cell(x1, y, x2, y, merged_cells, col_x, row_y):
                        cv2.line(display, (x1, y), (x2, y), (255, 0, 0), 2)

            # draw merged cell borders
            for (r1, c1, r2, c2) in merged_cells:
                x1, x2 = col_x[c1], col_x[c2 + 1]
                y1, y2 = row_y[r1], row_y[r2 + 1]
                cv2.rectangle(display, (x1, y1), (x2, y2), (0, 0, 255), 3)

            display_img = cv2.resize(display, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
            cv2.imshow("Merge cells", display_img)
            key = cv2.waitKey(20) & 0xFF

            if len(clicks) == 2:
                (x1, y1), (x2, y2) = clicks
                c1 = find_index(min(x1, x2), col_x)
                c2 = find_index(max(x1, x2), col_x)
                r1 = find_index(min(y1, y2), row_y)
                r2 = find_index(max(y1, y2), row_y)
                merged_cells.append((r1, c1, r2, c2))
                clicks.clear()

            if key in (8, ord('z')) and merged_cells:
                merged_cells.pop()

            if key == 27:  # ESC
                clicks.clear()

            if key == 13:  # ENTER
                break

        cv2.destroyAllWindows()

        def rotate_click(event, x, y, flags, param):
            if event not in (cv2.EVENT_LBUTTONDOWN, cv2.EVENT_RBUTTONDOWN):
                return

            orig_x = int(x / scale)
            orig_y = int(y / scale)

            c = find_index(orig_x, col_x)
            r = find_index(orig_y, row_y)

            merged = find_merged_region(r, c, merged_cells)

            # determine target region
            if merged:
                r1, c1, r2, c2 = merged
            else:
                r1 = r2 = r
                c1 = c2 = c

            if event == cv2.EVENT_LBUTTONDOWN:
                entry = (r1, c1, r2, c2, 90)
                if entry not in rotated_cells:
                    rotated_cells.append(entry)

            elif event == cv2.EVENT_RBUTTONDOWN:
                for i, (rr1, cc1, rr2, cc2, _) in enumerate(rotated_cells):
                    if rr1 == r1 and cc1 == c1 and rr2 == r2 and cc2 == c2:
                        rotated_cells.pop(i)
                        break

        cv2.namedWindow("Select rotated cells")
        cv2.setMouseCallback("Select rotated cells", rotate_click)

        while True:
            display = image.copy()

            # draw vertical grid lines (clipped)
            for i, x in enumerate(col_x):
                for r in range(len(row_y) - 1):
                    y1, y2 = row_y[r], row_y[r + 1]
                    if not intersects_merged_cell(x, y1, x, y2, merged_cells, col_x, row_y):
                        cv2.line(display, (x, y1), (x, y2), (0, 255, 0), 2)

            # draw horizontal grid lines (clipped)
            for i, y in enumerate(row_y):
                for c in range(len(col_x) - 1):
                    x1, x2 = col_x[c], col_x[c + 1]
                    if not intersects_merged_cell(x1, y, x2, y, merged_cells, col_x, row_y):
                        cv2.line(display, (x1, y), (x2, y), (255, 0, 0), 2)

            # draw merged cell borders
            for (r1, c1, r2, c2) in merged_cells:
                x1, x2 = col_x[c1], col_x[c2 + 1]
                y1, y2 = row_y[r1], row_y[r2 + 1]
                cv2.rectangle(display, (x1, y1), (x2, y2), (0, 0, 255), 3)

            # highlight rotated cells
            for r1, c1, r2, c2, _ in rotated_cells:
                x1, x2 = col_x[c1], col_x[c2 + 1]
                y1, y2 = row_y[r1], row_y[r2 + 1]
                cv2.rectangle(display, (x1, y1), (x2, y2), (255, 0, 255), 3)

            display_img = cv2.resize(display, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
            cv2.imshow("Select rotated cells", display_img)

            key = cv2.waitKey(20) & 0xFF
            if key == 13:  # ENTER
                break
            if key in (8, ord('z')) and rotated_cells:
                rotated_cells.pop()
            if key == 27:  # ESC
                rotated_cells.clear()

        cv2.destroyAllWindows()

    # ==============================
    # Save selection
    # ==============================
    save_selection(savefile, col_x, row_y, merged_cells, rotated_cells, image.shape[:2], dpi)
    print(f"\nSelection saved to {savefile}.")

    # ==============================
    # OCR over cells
    # ==============================
    rows = len(row_y) - 1
    cols = len(col_x) - 1
    data = [["" for _ in range(cols)] for _ in range(rows)]
    covered_cells = set()
    
    ocr_tasks = []

    rotated_lookup = {(r1, c1): angle for r1, c1, r2, c2, angle in rotated_cells}

    # 1. Process merged cells first
    for r1, c1, r2, c2 in merged_cells:
        if r1 < rows and c1 < cols and r2 < rows and c2 < cols:
            x1, x2 = col_x[c1], col_x[c2 + 1]
            y1, y2 = row_y[r1], row_y[r2 + 1]
            merged_img = gray[y1:y2, x1:x2]
            ocr_tasks.append((r1, c1, merged_img))
            
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    covered_cells.add((r, c))

    # 2. Process remaining cells
    for r in range(rows):
        for c in range(cols):
            if (r, c) not in covered_cells:
                x1, x2 = col_x[c], col_x[c + 1]
                y1, y2 = row_y[r], row_y[r + 1]
                ocr_tasks.append((r, c, gray[y1:y2, x1:x2]))

    # Execute OCR
    if concurrent_ocr:
        print(f"Starting concurrent OCR for {len(ocr_tasks)} cells...")
        with ThreadPoolExecutor() as executor:
            worker = partial(
                ocr_worker,
                ocr_lang=ocr_lang,
                rotated_lookup=rotated_lookup
            )
            # Map tasks to worker function
            with ThreadPoolExecutor() as executor:
                results = list(executor.map(worker, ocr_tasks))
            for r, c, text in results:
                data[r][c] = text
    else:
        for r, c, img in ocr_tasks:
            if (r, c) in rotated_lookup:
                img = rotate_for_ocr(img, rotated_lookup[(r, c)])
            data[r][c] = ocr_cell(img, lang=ocr_lang)

    # ==============================
    # Create DataFrame & Export (with layout)
    # ==============================
    df = pd.DataFrame(data)

    csv_path = f"output/{files[selection]}.csv"
    xlsx_path = f"output/{files[selection]}.xlsx"

    while True:
        try:
            df.to_csv(csv_path, index=False, header=False)
            df.to_excel(xlsx_path, index=False, header=False)

            # apply merged + rotated layout
            apply_excel_layout(
                xlsx_path,
                merged_cells,
                rotated_cells
            )

            print(f"\nTable extracted successfully to {csv_path} and {xlsx_path}\n")
            break

        except PermissionError:
            print("\nPermission denied while saving output files.")
            print("Please close the files in Excel.")

            retry = input("Retry saving? (y/n): ").strip().lower()
            if retry == "n":
                print("Saving aborted.")
                break
